import openpyxl
from openpyxl.styles import Font, Border, Alignment, PatternFill, Color, GradientFill, Side
from django.http import HttpResponse
from datetime import datetime
from django.utils.translation import gettext as _
from django.utils.timezone import localtime
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from collections import defaultdict
from operator import itemgetter



def export_grouped_report(modeladmin, request, queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reports Export"

    

    headers = [
    "No.", 
    "Offender Side",
    "Full Name of Offender",
    "Date of Birth",
    "Passport Number",
    "Citizenship",
    "Address",
    "Number of Offenders",
    "Customs Office",
    "Customs Checkpoint",
    "Explanatory Note Number",
    "Date of Preparation",
    "Goods",
    "Quantity",
    "Place of Discovery",
    "Entry / Exit / Transit",
    "Country of Origin",
    "Destination Country",
    "Vehicle Registration Number",
    "Articles of the Customs Code",
    "Imposed Fine (manat)",
    "Paid Fine (manat)",
    "Customs Officer"
]

    ws.append(headers)

    # Style headers
    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
        for cell in col:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

    # Step 1: Group reports by violation
    grouped_data = []
    seen_reports = set()

    violations = queryset.values_list('violation', flat=True).distinct()

    for violation_id in violations:
        related_reports = queryset.filter(violation__id=violation_id).select_related(
            'violation', 'customspoint', 'customspoint__customsoffice', 'customsofficer'
        ).prefetch_related('stored_goods__product', 'stored_goods__unitofmeasurement', 'administration_codexes', 'assigned_tasks')

        if not related_reports.exists():
            continue

        # Avoid duplicate entries
        unique_reports = [r for r in related_reports if r.pk not in seen_reports]
        if not unique_reports:
            continue

        violation = unique_reports[0].violation
        violation_type = violation.violation_type or ''

        if violation_type.lower() == "legal entity":
            fullname = f"{violation.company_name or ''}\nÝolbaşçysy: {violation.company_boss_fullname or ''}"
            address = violation.address
        else:
            fullname = f"{violation.violator_surname or ''} {violation.violator_name or ''} {violation.father_name or ''}".strip()
            address = violation.violator_address

        dob = violation.date_of_birth.strftime('%d.%m.%Y') if violation.date_of_birth else ''
        passport = violation.passport_number or ''
        nationality = getattr(violation.nationality, 'name', '') if hasattr(violation, 'nationality') else ''
        report_count = len(unique_reports)

        grouped_data.append({
            "violation_type": violation_type,
            "fullname": fullname,
            "dob": dob,
            "passport": passport,
            "nationality": nationality,
            "address": address,
            "count": report_count,
            "reports": unique_reports
        })

        # Mark reports as seen
        seen_reports.update(r.pk for r in unique_reports)

    # Step 2: Sort by count descending
    grouped_data.sort(key=itemgetter('count'), reverse=True)

    # Step 3: Write data to sheet
    row_num = 2
    tb_index = 1

    for group in grouped_data:
        first_row_written = False
        for report in group['reports']:
            stored_goods = report.stored_goods.all()
            if stored_goods:
                for sg in stored_goods:
                    row_data = [
                        tb_index if not first_row_written else '',
                        group['violation_type'] if not first_row_written else '',
                        group['fullname'] if not first_row_written else '',
                        group['dob'] if not first_row_written else '',
                        group['passport'] if not first_row_written else '',
                        group['nationality'] if not first_row_written else '',
                        group['address'] if not first_row_written else '',
                        group['count'] if not first_row_written else '',
                        report.customspoint.customsoffice.name if report.customspoint and report.customspoint.customsoffice else '',
                        report.customspoint.name if report.customspoint else '',
                        report.protocol_number,
                        report.report_date.strftime('%d.%m.%Y') if report.report_date else '',
                        sg.product.name if sg.product else '',
                        f"{sg.amount} {sg.unitofmeasurement.name if sg.unitofmeasurement else ''}",
                        str(sg.reasonforruleviolation) if sg.reasonforruleviolation else '',
                        report.entry_exit_transit,
                        str(report.from_country) if report.from_country else '',
                        str(report.to_country) if report.to_country else '',
                        str(report.carnumber) if report.carnumber else '',
                        ", ".join([str(c) for c in report.administration_codexes.all()]),
                        report.assigned_tasks.first().salnan_jerime if report.assigned_tasks.exists() else '',
                        report.assigned_tasks.first().tolenen_manat if report.assigned_tasks.exists() else '',
                        f"{report.customsofficer.surname} {report.customsofficer.name} {report.customsofficer.midname}".strip() if report.customsofficer else ''
                    ]
                    for col_index, value in enumerate(row_data, start=1):
                        cell = ws.cell(row=row_num, column=col_index, value=value)
                        if col_index in [3, 7]:  # wrap long text columns
                            cell.alignment = Alignment(wrap_text=True)

                    row_num += 1
                    first_row_written = True
            else:
                row_data = [
                    tb_index if not first_row_written else '',
                    group['violation_type'] if not first_row_written else '',
                    group['fullname'] if not first_row_written else '',
                    group['dob'] if not first_row_written else '',
                    group['passport'] if not first_row_written else '',
                    group['nationality'] if not first_row_written else '',
                    group['address'] if not first_row_written else '',
                    group['count'] if not first_row_written else '',
                    report.customspoint.customsoffice.name if report.customspoint and report.customspoint.customsoffice else '',
                    report.customspoint.name if report.customspoint else '',
                    report.protocol_number,
                    report.report_date.strftime('%d.%m.%Y') if report.report_date else '',
                    '', '', '',  # No stored goods
                    report.entry_exit_transit,
                    str(report.from_country) if report.from_country else '',
                    str(report.to_country) if report.to_country else '',
                    str(report.carnumber) if report.carnumber else '',
                    ", ".join([str(c) for c in report.administration_codexes.all()]),
                    report.assigned_tasks.first().salnan_jerime if report.assigned_tasks.exists() else '',
                    report.assigned_tasks.first().tolenen_manat if report.assigned_tasks.exists() else '',
                    f"{report.customsofficer.surname} {report.customsofficer.name} {report.customsofficer.midname}".strip() if report.customsofficer else ''
                ]
                for col_index, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_num, column=col_index, value=value)
                    if col_index in [3, 7]:
                        cell.alignment = Alignment(wrap_text=True)

                row_num += 1
                first_row_written = True

        tb_index += 1

    # Adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['G'].width = 35

    # Return Excel file as HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"report_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

export_grouped_report.short_description = _('1+ saklananlar')

def export_report_to_excel(self, request, queryset):
        # Create an Excel workbook and a sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = str(_('Reports'))

        # Define the translated headers, including military name and position
        headers = [
            str(_('T/b')),
            str(_('Ish Toplum Number')),
            str(_('Protocol Number')),
            str(_('Report date')),
            str(_('Customs Office')),
            str(_('Customs Point')),
            str(_('Violation Type')),
            str(_('Violation')),
            str(_('Passport Number')),
            str(_('Passport Issue Date')),
            str(_('Date of Birth')),
            str(_('Place of Birth')),
            str(_('Address')),
            str(_('Phone')),
            str(_('Nationality')),
            str(_('Product Counter')),
            str(_('Product Name')),
            str(_('Amount')),
            str(_('Unit of Measurement')),
            str(_('Unit of Measurement')),
            str(_('Reason for Rule Violation')), # Yuze cykarylan yeri
            str(_('Entry/Exit/Transit')),
            str(_('From Country')),
            str(_('To Country')),
            str(_('Vehicle Brand')),
            str(_('Car Number')),
            str(_('Transport Company Name')),
            str(_('Basis for Discovery')),
            str(_('Method of Discovery')),
            # str(_('Reason for Rule Violation')),
            str(_('Administration Codexes')),

            str(_('Customs Officer')),
            str(_('Position')),
            str(_('Military Name')),
            str(_('Language of Work Conducted')),
            str(_('Witness Full Names')),   
            str(_('Witness Address')),
              
            # Assigned Task fields
            str(_('Assigned Task Manat')),
            str(_('Assigned Task Workgroup')),

            # Assigned Letter fields
            str(_('Assigned Letter Number')),
            str(_('Assigned Letter Date')), 
            str(_('Letter for Action')),

            str(_('Created At')),
            str(_('Updated At')),  
                     
        ]
        sheet.append(headers)

        # Set header styles
        header_font = Font(bold=True, color='FFFFFF', name='Times New Roman', size=12)
        header_fill = GradientFill(stop=[Color(rgb='0072B2'), Color(rgb='0094D8')], type='linear', degree=0)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
       
        
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        sheet.row_dimensions[1].height = 45  # Set header row height
    
        # Initialize max_lengths for each column
        max_lengths = [len(header) for header in headers]  # Start with header lengths


          # Set font for the data rows
        data_font = Font(name='Times New Roman', size=12)

        # Utility function to safely return string values
        def safe_str(value):
            return str(value) if value is not None else ''

        # Utility function to construct the full name
        def get_full_name(violator_name, violator_surname, father_name):
            return ' '.join(filter(None, [violator_name, violator_surname, father_name]))

        # Write the translated data for each report
        for index, report in enumerate(queryset, start=1):

            # Determine row fill color for alternating rows
            if index % 2 == 0:
                row_fill = PatternFill(start_color='EAEAEA', end_color='EAEAEA', fill_type='solid')
            else:
                row_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

              
            # Format administration codexes
            administration_codexes = ', '.join(report.administration_codexes.values_list('name', flat=True))

            # Convert timezone-aware datetimes to naive datetimes
            created_at_naive = localtime(report.created_at).replace(tzinfo=None)
            updated_at_naive = localtime(report.updated_at).replace(tzinfo=None)

            # Format the report date in 'd.m.y' format
            report_date_formatted = report.report_date.strftime('%d.%m.%Y') if report.report_date else ''
          
            # Extract CustomsOfficer info (name, surname, position, military name)
            if report.customsofficer:
                officer_full_name = f"{report.customsofficer.name} {report.customsofficer.surname}".strip()
                officer_position = safe_str(report.customsofficer.position.name) if report.customsofficer.position else ''
                officer_militaryname = safe_str(report.customsofficer.militaryname.name) if report.customsofficer.militaryname else ''
            else:
                officer_full_name = ''
                officer_position = ''
                officer_militaryname = ''

            # Extract Violation info (violation_type, full name for individuals, company for legal entities)
            if report.violation:
                violation_type = safe_str(report.violation.violation_type)

                # Handle full name for individuals (individual or official)
                if violation_type in ['individual', 'official']:
                    violator_full_name = get_full_name(
                        report.violation.violator_name,
                        report.violation.violator_surname,
                        report.violation.father_name
                    )
                    company_name = ''
                # Handle company name for legal entities (legal entity)
                elif violation_type == 'legal entity':
                    violator_full_name = ''
                    company_name = safe_str(report.violation.company_name)
                else:
                    violator_full_name = ''
                    company_name = ''
                    
                # Safely handle the address
                violator_address = safe_str(report.violation.violator_address or report.violation.address)
                violator_phone = safe_str(report.violation.phone)
                violator_date_of_birth = report.violation.date_of_birth.strftime('%d.%m.%Y') if report.violation.date_of_birth else ''
                violator_place_of_birth = safe_str(report.violation.place_of_birth)
                violator_passport_number = safe_str(report.violation.passport_number)
                violator_passport_issue_date = report.violation.passport_issue_date.strftime('%d.%m.%Y') if report.violation.passport_issue_date else ''
                violator_nationality = safe_str(report.violation.nationality.name if report.violation.nationality else '')

            else:
                violation_type = ''
                violator_full_name = ''
                company_name = ''
                violator_address = ''
                violator_phone = ''
                violator_date_of_birth = ''
                violator_place_of_birth = ''
                violator_passport_number = ''
                violator_passport_issue_date = ''
                violator_nationality = ''
            
            # Extract witness data
            # witnesses_full_names = ', '.join([witness.fullname for witness in report.witnesses.all()])
            # witnesses_addresses = ', '.join([witness.address for witness in report.witnesses.all()])
            witnesses_full_names = ', '.join([witness.fullname if witness.fullname else ' ' for witness in report.witnesses.all()])
            witnesses_addresses = ', '.join([witness.address if witness.address else ' ' for witness in report.witnesses.all()])

            # Extract assigned task data
            assigned_tasks_manat = ', '.join([f"{task.salnan_jerime}" for task in report.assigned_tasks.all()])
            assigned_tasks_workgroups = ', '.join([f"{task.workgroup.name}" for task in report.assigned_tasks.all()])
            
            # Extract assigned letter data related to each task
            letter_numbers = ', '.join([letter.number for task in report.assigned_tasks.all() for letter in task.assigned_letters.all()])
            letter_dates = ', '.join([letter.date.strftime('%d.%m.%Y') for task in report.assigned_tasks.all() for letter in task.assigned_letters.all()])
            
            # Extract letter for action names related to each assigned letter
            letter_for_actions = ', '.join([safe_str(letter.letterforaction.name) for task in report.assigned_tasks.all() for letter in task.assigned_letters.all()])

            # Build the translated row
            # Group rows under a single index
            products = report.stored_goods.all() 
            if not products:
                # If there are no stored goods, still generate one row for the report, with empty stored good columns.
                products = [None] 
            for idx, product in enumerate(products):
                row = [
                    index if idx == 0 else '',  # Show index only for the first row of the group
                    report.ish_toplum_number ,  
                    report.protocol_number ,
                    report_date_formatted,
                    report.customsoffice.name ,
                    report.customspoint.name ,
                    violation_type ,
                    violator_full_name if violation_type in ['individual', 'official'] else (
                        company_name if violation_type == 'legal entity' else ''
                    ),
                    violator_passport_number ,
                    violator_passport_issue_date ,
                    violator_date_of_birth ,
                    violator_place_of_birth ,
                    violator_address ,
                    violator_phone ,
                    violator_nationality ,
                    
                    idx + 1,
                    product.product.name if product else '',  # Product name (check if product exists)
                    product.amount if product else '',        # Amount
                    product.unitofmeasurement.name if product and product.unitofmeasurement else '',  # Unit of measurement
                    product.reasonforruleviolation.name if product and product.reasonforruleviolation else '',
                    product.note if product else '',  

                    str(_(report.entry_exit_transit)) ,
                    report.from_country.name if  report.from_country else '',
                    report.to_country.name if  report.to_country else '',
                    report.vehicle_brand.name if   report.vehicle_brand else '',
                    report.carnumber,
                    report.transport_company.name if  report.transport_company else '',
                    report.basisfordiscovery.name if report.basisfordiscovery else '',
                    report.methodofdiscovery.name if report.methodofdiscovery else '',
                    # report.reasonforruleviolation.name if idx == 0 and report.reasonforruleviolation else '',
                    administration_codexes ,
                    officer_full_name ,
                    officer_position ,
                    officer_militaryname ,
                    report.language_of_work_conducted ,
                    witnesses_full_names ,
                    witnesses_addresses ,

                    # Include Assigned Task data
                    assigned_tasks_manat ,
                    assigned_tasks_workgroups ,
                   

                    # Include Assigned Letter data
                    letter_numbers ,
                    letter_dates ,
                    letter_for_actions ,

                    created_at_naive ,
                    updated_at_naive ,
                ]
                sheet.append(row)
                # Update max_lengths based on the current row's data
                for i, item in enumerate(row):
                    max_lengths[i] = max(max_lengths[i], len(safe_str(item)))
                
                # Get the last row added
                last_row = sheet.max_row
                
                # Set font and height for each data row
                for cell in sheet[last_row]:  # +1 for header row
                    if violation_type:  # Condition for violations
                        cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red fill for violations
                    else:
                        cell.fill = row_fill 
                    cell.fill = row_fill 
                    cell.font = data_font
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    

        sheet.row_dimensions[last_row].height = 25  
        

            

            # Set the column widths based on the max lengths
        for i, max_length in enumerate(max_lengths):
            sheet.column_dimensions[get_column_letter(i + 1)].width = 20


        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        # Add autofilter to the headers
        sheet.auto_filter.ref = sheet.dimensions

        # Freeze header row for better usability
        sheet.freeze_panes = 'A3'
        sheet.column_dimensions['A'].width = 5
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['AJ'].width = 35


        # Prepare the response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        workbook.save(response)
        return response

export_report_to_excel.short_description = _('Export selected reports to Excel')