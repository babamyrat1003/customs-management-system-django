from django import forms
from django.contrib import admin
from admin_searchable_dropdown.filters import AutocompleteFilter
import nested_admin
from .forms import CustomsPointForm, MethodOfDiscoveryForm, ViolationAdminForm
from .models import TJK, City, Country, DernewGornush, DernewNetijesi,ProductCategory, TransportCompanyName, VehicleBrand, CustomsOffice, CustomsPoint, LettersForAction, Product, StoredGood, StoredGoodImage, UnitOfMeasurement, MethodOfDiscovery, ReasonForRuleViolation, BasisForDiscovery, Violation, Workgroup, MilitaryName, Position, AdministrationCodex, CustomsOfficer
from django.utils import translation
from django.utils.translation import gettext_lazy as _ , get_language
from django.http import HttpResponse
from django.utils import timezone
from django.utils import formats
from django.utils.html import format_html
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, PatternFill, Font, Side
import openpyxl


class CustomsPointInline(admin.TabularInline):  # or admin.StackedInline
    model = CustomsPoint
    extra = 1  # Number of empty forms to display by default


@admin.register(CustomsOffice)
class CustomsOfficeAdmin(admin.ModelAdmin):
    # Fields to display in the list view
    list_display = ('name','code','created_at', 'updated_at')
    
    # Fields to make searchable
    search_fields = ('name','code',)
    
    # Filters for the sidebar
    list_filter = ('created_at',)
    
    # Date-based navigation
    date_hierarchy = 'created_at'
    
    # Ordering by default
    ordering = ('name',)
    
    # Read-only fields to prevent editing
    readonly_fields = ('created_at', 'updated_at')
    
    
    # Fields to be displayed in the form
    fieldsets = (
        (None, {
            'fields': ('name','code',)
        }),

    )
    
    # Add bulk actions (including the Excel export)
    actions = ['export_to_csv', 'export_to_excel']  # Include both actions



   # Custom admin action: Example to export data as Excel
    def export_to_excel(self, request, queryset):
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        border = Border(left=Side(border_style="thin"), 
                        right=Side(border_style="thin"), 
                        top=Side(border_style="thin"), 
                        bottom=Side(border_style="thin"))

        # Get the current language
        current_language = translation.get_language()
        
        # Define headers based on the current language
        if current_language == 'tk':  # Turkmen
            headers = ['Name', 'Created At', 'Updated At']
        else:  # Default to English
            headers = ['Name', 'Created At', 'Updated At']
        
        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Customs Offices"
        
        # Write the headers to the first row with formatting
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        
        # Write the data rows
        for office in queryset:
            created_at_naive = office.created_at.astimezone(timezone.get_current_timezone()).replace(tzinfo=None)
            updated_at_naive = office.updated_at.astimezone(timezone.get_current_timezone()).replace(tzinfo=None)
            
            ws.append([office.name, created_at_naive, updated_at_naive])
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Create the HTTP response with the appropriate Excel content type
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="customs_offices_{current_language}.xlsx"'
        
        # Save the workbook to the response
        wb.save(response)
        
        return response

    export_to_excel.short_description = _("Export selected customs offices to Excel")

class CustomsOfficeFilter(AutocompleteFilter):
    title = _('CustomsOffice') # display title
    field_name = 'customsoffice' # name of the foreign key field
    
def export_to_excel_customs_point(modeladmin, request, queryset):
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border = Border(left=Side(border_style="thin"), 
                    right=Side(border_style="thin"), 
                    top=Side(border_style="thin"), 
                    bottom=Side(border_style="thin"))

    # Get the current language
    current_language = get_language()
    
    # Define headers based on the current language
    if current_language == 'tk':  # Turkmen
        headers = [str(_('Ady')), str(_('Customs Office')), str(_('Created At')), str(_('Updated At'))]
    else:  # Default to English
        headers = [str(_('Name')), str(_('Customs Office')), str(_('Created At')), str(_('Updated At'))]

    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(_("Customs Points"))

    # Write the headers to the first row with formatting
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # Write the data rows
    for point in queryset:
        created_at_formatted = formats.date_format(point.created_at, "DATETIME_FORMAT", use_l10n=True)
        updated_at_formatted = formats.date_format(point.updated_at, "DATETIME_FORMAT", use_l10n=True)
        
        ws.append([
            point.name,
            point.customsoffice.name,
            created_at_formatted,
            updated_at_formatted,
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except (AttributeError, TypeError) as e:
                # Handle specific exceptions or ignore
                print(f"Skipping cell due to error: {e}")
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Create the HTTP response with the appropriate Excel content type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="customs_points_{current_language}.xlsx"'

    # Save the workbook to the response
    wb.save(response)

    return response

export_to_excel_customs_point.short_description = _("Export selected customs points to Excel")

@admin.register(CustomsPoint)
class CustomsPointAdmin(admin.ModelAdmin):
    form = CustomsPointForm
    list_display = ('name', 'code','customsoffice', 'created_at', 'updated_at')
    search_fields = ('name', 'code','customsoffice__name')
    list_filter = (CustomsOfficeFilter,'created_at',)
    autocomplete_fields = ['customsoffice'] 
    list_per_page = 15
    prepopulated_fields = {'name': ('customsoffice',)}
    actions = [export_to_excel_customs_point]  # Add the export action

class CityInline(admin.TabularInline):  # or admin.StackedInline
    model = City
    extra = 1  # Number of empty forms to display

@admin.register(Country)
class CountryAdmin(admin.ModelAdmin):
    list_display = ('name', 'code', 'created_at', 'updated_at')
    search_fields = ('name', 'code')
    list_filter = ('created_at',)
    ordering = ('name',)
    readonly_fields = ('created_at', 'updated_at')
    list_per_page = 15
    inlines = [CityInline] 

    fieldsets = (
        (None, {
            'fields': ('name', 'code'),
            'description': _('Manage countries and their respective codes.'),
        }),

    )

@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ('name', 'productcategory','created_at', 'updated_at')
    search_fields = ('name',)
    list_filter = ('created_at','productcategory')
    ordering = ('name',)
    readonly_fields = ('created_at', 'updated_at')
    list_per_page = 16
    list_max_show_all = 1000  # Set this to the maximum number of items you want to show in "Show all"
    show_full_result_count = True

    fieldsets = (
        (None, {
            'fields': ('productcategory','name',),
            'description': _('Manage products'),
        }),

    )

    
@admin.register(UnitOfMeasurement)
class UnitOfMeasurementAdmin(admin.ModelAdmin):
    list_display = ('name', 'description', 'created_at', 'updated_at')  # Display important fields
    search_fields = ('name', 'description')  # Allow searching by name and description
    list_filter = ('created_at', 'updated_at')  # Filter by created and updated timestamps
    readonly_fields = ('created_at', 'updated_at')  # Make timestamps read-only
    list_per_page = 15  # Set pagination for list view

    fieldsets = (
        (None, {
            'fields': ('name', 'description'),
            'description': _('Manage unit of measurements'),
        }),

    )

@admin.register(MethodOfDiscovery)
class MethodOfDiscoveryAdmin(admin.ModelAdmin):
    form = MethodOfDiscoveryForm
    list_display = ('name', 'description', 'created_at', 'updated_at')  # Display important fields
    search_fields = ('name', 'description')  # Allow searching by name and description
    list_filter = ('created_at', 'updated_at')  # Filter by created and updated timestamps
    readonly_fields = ('created_at', 'updated_at')  # Make timestamps read-only
    list_per_page = 20  # Set pagination for list view

    fieldsets = (
        (None, {
            'fields': ('name', 'description'),
            'description': _('Manage Method of Discoveries'),
        }),

    )

@admin.register(ReasonForRuleViolation)
class ReasonForRuleViolationAdmin(admin.ModelAdmin):
    form = MethodOfDiscoveryForm
    list_display = ('name', 'description', 'created_at', 'updated_at')  # Display important fields
    search_fields = ('name', 'description')  # Allow searching by name and description
    list_filter = ('created_at', 'updated_at')  # Filter by created and updated timestamps
    readonly_fields = ('created_at', 'updated_at')  # Make timestamps read-only
    list_per_page = 15  # Set pagination for list view

    fieldsets = (
        (None, {
            'fields': ('name', 'description'),
            'description': _('Manage Reason for Rule Violation'),
        }),

    )

@admin.register(BasisForDiscovery)
class BasisForDiscoveryAdmin(admin.ModelAdmin):
    list_display = ('name', 'description', 'created_at', 'updated_at')  # Display important fields
    search_fields = ('name', 'description')  # Allow searching by name and description
    list_filter = ('created_at', 'updated_at')  # Filter by created and updated timestamps
    readonly_fields = ('created_at', 'updated_at')  # Make timestamps read-only
    list_per_page = 15  # Set pagination for list view

    fieldsets = (
        (None, {
            'fields': ('name', 'description'),
            'description': _('Manage Reason for Rule Violation'),
        }),

    )

@admin.register(Workgroup)
class WorkgroupAdmin(admin.ModelAdmin):
    list_display = ('name', 'description', 'created_at', 'updated_at')  # Display important fields
    search_fields = ('name', 'description')  # Allow searching by name and description
    list_filter = ('created_at', 'updated_at')  # Filter by created and updated timestamps
    readonly_fields = ('created_at', 'updated_at')  # Make timestamps read-only
    list_per_page = 15  # Set pagination for list view

    fieldsets = (
        (None, {
            'fields': ('name', 'description'),
            'description': _('Manage Reason for Rule Violation'),
        }),

    )

@admin.register(LettersForAction)
class LettersForActionAdmin(admin.ModelAdmin):
    list_display = ('name', 'description', 'created_at', 'updated_at')  # Display important fields
    search_fields = ('name', 'description')  # Allow searching by name and description
    list_filter = ('created_at', 'updated_at')  # Filter by created and updated timestamps
    readonly_fields = ('created_at', 'updated_at')  # Make timestamps read-only
    list_per_page = 15  # Set pagination for list view

    fieldsets = (
        (None, {
            'fields': ('name', 'description'),
            'description': _('Manage Reason for Rule Violation'),
        }),

    )

@admin.register(MilitaryName)
class MilitaryNameAdmin(admin.ModelAdmin):
    list_display = ('name', 'created_at', 'updated_at')  # Display important fields
    search_fields = ('name',)  # Allow searching by name and description
    list_filter = ('created_at', 'updated_at')  # Filter by created and updated timestamps
    readonly_fields = ('created_at', 'updated_at')  # Make timestamps read-only
    list_per_page = 15  # Set pagination for list view
    ordering = ('created_at',)

    fieldsets = (
        (None, {
            'fields': ('name',),
            'description': _('Manage Military Names'),
        }),

    )

@admin.register(Position)
class PositionAdmin(admin.ModelAdmin):
    list_display = ('name', 'created_at', 'updated_at')  # Display important fields
    search_fields = ('name',)  # Allow searching by name and description
    list_filter = ('created_at', 'updated_at')  # Filter by created and updated timestamps
    readonly_fields = ('created_at', 'updated_at')  # Make timestamps read-only
    list_per_page = 15  # Set pagination for list view
    ordering = ('created_at',)

    fieldsets = (
        (None, {
            'fields': ('name',),
            'description': _('Manage Military Names'),
        }),
      

    )

@admin.register(AdministrationCodex)
class AdministrationCodexAdmin(admin.ModelAdmin):
    list_display = ('name', 'description', 'created_at', 'updated_at')  # Display important fields
    search_fields = ('name', 'description')  # Allow searching by name and description
    list_filter = ('created_at', 'updated_at')  # Filter by created and updated timestamps
    readonly_fields = ('created_at', 'updated_at')  # Make timestamps read-only
    list_per_page = 7  # Set pagination for list view

    fieldsets = (
        (None, {
            'fields': ('name', 'description'),
            'description': _('Manage Reason for Rule Violation'),
        }),

    )
    
@admin.register(CustomsOfficer)
class CustomsOfficerAdmin(admin.ModelAdmin):

    list_display = ('full_name', 'position', 'militaryname', 'created_at', 'updated_at')

    list_filter = ('position', 'militaryname')

    search_fields = ('name', 'surname', 'midname', 'position__name', 'militaryname__name')

    autocomplete_fields = ['position', 'militaryname']

    readonly_fields = ('created_at', 'updated_at')

    ordering = ['surname', 'name']

    list_per_page = 20
    fieldsets = (
        ('Personal Information', {
            'fields': ('name', 'surname', 'midname')
        }),
        ('Professional Information', {
            'fields': ('position', 'militaryname')
        }),

    )

    # Custom method to display the full name of the customs officer
    def full_name(self, obj):
        """Display the full name of the customs officer."""
        return f"{obj.surname} {obj.name} {obj.midname}".strip()
    
    full_name.short_description = _('Full Name')  # Set display name in the admin interface

class CitizenshipFilter(AutocompleteFilter):
    title = _('Raýatlylygy')
    field_name = 'nationality' 

class ViolationAdmin(admin.ModelAdmin):
    form = ViolationAdminForm
    list_display = ('violation_type', 'full_name', 'company_name','company_boss_fullname','nationality', 'passport_number','date_of_birth','place_of_birth','created_at','updated_at')
    search_fields = ['violation_type', 'passport_number','violator_name', 'violator_surname','father_name', 'company_name','company_boss_fullname', 'nationality__name','nationality__code']
    autocomplete_fields = ['nationality']
    # readonly_fields = ('created_at', 'updated_at')
    list_filter = ('violation_type', CitizenshipFilter)
    list_per_page = 15
    exclude = ('created_at', 'updated_at')

    def full_name(self, obj):
        return obj.full_name
    full_name.short_description = _('Full Name')
    
    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if obj:
            form.instance = obj 
        return form
    
    
    class Media:
        js = ('js/admin.js',) 

admin.site.register(Violation, ViolationAdmin)


class StoredGoodImageInline(nested_admin.NestedStackedInline):
    model = StoredGoodImage
    fields = ('image', 'description', 'image_preview')
    readonly_fields = ('image_preview',) 
    extra = 1
    verbose_name = _("Stored Good Image")
    verbose_name_plural = _("Stored Good Images")

    def get_extra(self, request, obj=None, **kwargs):
        # Check if the object is being updated
        if obj:
            return 0  # No extra rows when updating
        return 1  # Add one extra row when creating a new Report
    

    def image_preview(self, obj):
        """
        Display a clickable image preview in the admin interface. Clicking the image will open it in full size.
        """
        if obj.image:  # Check if an image is available
            # Create a clickable thumbnail that opens the image in full size
            return format_html(
                '<a href="{0}" target="_blank">'
                '<img src="{0}" style="max-height: 200px; max-width: 200px;" />'
                '</a>',
                obj.image.url
            )
        return "No image available"

    image_preview.short_description = 'Image Preview'

class StoredGoodInline(nested_admin.NestedTabularInline):
    model = StoredGood 
    fields = ('product', 'amount', 'unitofmeasurement', 'reasonforruleviolation','note')
    extra = 1
    verbose_name = _("Stored Good")
    verbose_name_plural = _("Stored Goods")
    autocomplete_fields = ['product']
    inlines = [StoredGoodImageInline]

    def get_extra(self, request, obj=None, **kwargs):
        if obj:
            return 0  
        return 1 


class StoredGoodAdmin(admin.ModelAdmin):
    inlines = [StoredGoodImageInline]
    list_display = ('product', 'amount', 'unitofmeasurement', 'created_at', 'updated_at')
    search_fields = ('product__name', 'unitofmeasurement__name', 'note')
    autocomplete_fields = ['product']
    list_filter = ('product', 'unitofmeasurement', 'created_at')
    readonly_fields = ('created_at', 'updated_at')
    ordering = ['-created_at']
    
    fieldsets = (
        (None, {
            'fields': ('product', 'amount', 'unitofmeasurement')
        }),
        (_('Additional Information'), {
            'fields': ('note',),            
        }),

    )

# admin.site.register(StoredGood, StoredGoodAdmin)


@admin.register(VehicleBrand)
class VehicleBrandAdmin(admin.ModelAdmin):
    list_display = ('name',)
    search_fields = ('name',)


@admin.register(TransportCompanyName)
class TransportCompanyNameAdmin(admin.ModelAdmin):
    list_display = ('name',) 
    search_fields = ('name',) 

@admin.register(ProductCategory)
class ProductCategoryAdmin(admin.ModelAdmin):
    list_display = ('name',)
    search_fields = ('name',)

@admin.register(DernewGornush)
class DernewGornushAdmin(admin.ModelAdmin):
    list_display = ('name',)
    search_fields = ('name',)

@admin.register(TJK)
class TJKAdmin(admin.ModelAdmin):
    list_display = ('name', 'description',)  
    search_fields = ('name', 'description') 
    list_per_page = 7




class DernewNetijesiInline(nested_admin.NestedStackedInline):
    model = DernewNetijesi 
    fields = ('dernew', 'tjk', 'workgroup', 'gorlen_care','hatyn_nusgasy','hatyn_nusgasy_link')
    extra = 1
    readonly_fields = ('hatyn_nusgasy_link',)
    verbose_name = _("Derňew netijesi")
    verbose_name_plural = _("Derňew netijeler")
    autocomplete_fields = ['dernew','tjk', 'workgroup',]

    def hatyn_nusgasy_link(self, obj):
        """Show a link to download the uploaded PDF file."""
        if obj.hatyn_nusgasy:
            return format_html('<a href="{}" target="_blank">📄 Download PDF</a>', obj.hatyn_nusgasy.url)
        return "No file uploaded"
    
    hatyn_nusgasy_link.short_description = "Hatyň nusgasy PDF"

    def get_extra(self, request, obj=None, **kwargs):
        if obj:
            return 0  
        return 1 
    
@admin.register(DernewNetijesi)
class DernewNetijesiAdmin(admin.ModelAdmin):
    list_display = ('dernew', 'tjk', 'workgroup', 'gorlen_care', 'hatyn_nusgasy_link')  # Display key fields
    list_filter = ('dernew', 'tjk', 'workgroup')  # Add filtering
    search_fields = ('dernew__name', 'tjk__name', 'workgroup__name')  # Enable search
    autocomplete_fields = ('dernew', 'tjk', 'workgroup')  # Optimize performance
    readonly_fields = ('hatyn_nusgasy_link',)  # Prevent unnecessary edits

    fieldsets = (
        (_("General Information"), {
            'fields': ('dernew', 'tjk', 'workgroup')
        }),
        (_("Details"), {
            'fields': ('gorlen_care', 'hatyn_nusgasy', 'hatyn_nusgasy_link'),
        }),
    )

    def hatyn_nusgasy_link(self, obj):
        """Show a link to download the uploaded PDF file."""
        if obj.hatyn_nusgasy:
            return format_html('<a href="{}" target="_blank">📄 Download PDF</a>', obj.hatyn_nusgasy.url)
        return "No file uploaded"
    
    hatyn_nusgasy_link.short_description = "Copy of the letter PDF"
